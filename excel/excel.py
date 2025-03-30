from openpyxl import load_workbook

class ArchivoExcel:

    def __init__(self, archivo):
        self.wb=load_workbook(filename=archivo)
        self.nombre=archivo

    def obtenerDatos(self,fila,columna):
        ws=self.wb["Hoja1"]
        valor=ws.cell(row=fila, column=columna)
        return valor.value

    def modificarValor(self,fila,columna,valor):
        ws = self.wb["Hoja1"]
        cell=ws.cell(row=fila, column=columna)
        cell.value=valor
        self.wb.save(self.nombre)











if __name__ == '__main__':
    archivo=ArchivoExcel("unidades.xlsx")
    print(archivo.obtenerDatos(1,1))








